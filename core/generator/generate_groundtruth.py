from core.generator.generate_excel_utils import (
    create_empty_excel_with_sheets,
    save_excel_workbook,
    extract_cve_sort_key,
    apply_hyperlinks,
    apply_ranking_colors
)
from data.models import CAPEC
from threatlinker.paths import REPORTS
import traceback
from openpyxl.styles import PatternFill, Border, Side
from core.metrics.metrics_functions import mean_reciprocal_rank, recall_at_k

def create_groundtruth_excel(task, groundtruth, ai_models, top_count):
    """
    Genera un file Excel per i risultati di una GroundTruth e la Task associata.

    :param task: Oggetto Task associato.
    :param groundtruth: Oggetto GroundTruth selezionato.
    :param ai_models: Lista dei modelli AI da considerare.
    :return: Percorso del file Excel generato.
    """
    print(f"Starting create_groundtruth_excel for Task ID: {task.id}, GroundTruth ID: {groundtruth.id}")

    # Crea un workbook con due fogli: "results" e "stats"
    sheet_names = ["results", "stats"]
    workbook, sheet_objects = create_empty_excel_with_sheets(sheet_names)
    results_sheet = sheet_objects["results"]
    stats_sheet = sheet_objects["stats"]

    # Prendi la mapping di GroundTruth
    groundtruth_mapping = groundtruth.mapping  # {'CVE-XXX': ['CAPEC-1', 'CAPEC-2']}
    print(f"GroundTruth Mapping: {groundtruth_mapping}")

    # Recupera tutte le SingleCorrelations della Task
    single_correlations = task.single_correlations.all()
    print(f"Total SingleCorrelations: {single_correlations.count()}")

    # Colonne aggiuntive per i modelli AI
    model_columns = [f"{model}_Rank" for model in ai_models]

    # Scrive l'intestazione
    headers = ["CVE_ID", "CAPEC_ID", "CAPEC_Name"] + model_columns
    results_sheet.append(headers)
    print(f"Headers: {headers}")

    # Prendi tutte le CVE e le CAPEC associate da SingleCorrelation e GroundTruth
    rows = []
    print("Starting processing SingleCorrelations...")
    for single_correlation in single_correlations:
        cve_id = single_correlation.cve_id
        print(f"Processing SingleCorrelation: {single_correlation}, CVE ID: {cve_id}")

        if cve_id in groundtruth_mapping:
            capecs = groundtruth_mapping[cve_id]  # Lista di CAPEC associate a questa CVE
            print(f"CAPECs for CVE {cve_id}: {capecs}")

            for capec_id in capecs:
                try:
                    capec = CAPEC.objects.get(id=capec_id)
                    capec_name = capec.name
                    print(f"CAPEC Found: ID={capec_id}, Name={capec_name}")
                except CAPEC.DoesNotExist:
                    capec_name = "Unknown CAPEC"
                    print(f"CAPEC not found in database: {capec_id}")

                # Recupera i rank dai similarity_scores per ogni modello AI
                rank_data = []
                for model in ai_models:
                    similarity_scores = single_correlation.similarity_scores.get(model, [])
                    print(f"Similarity Scores for Model {model}: {similarity_scores}")

                    # Trova il rank per questa CAPEC
                    model_rank = None
                    for score_data in similarity_scores:
                        if score_data[0] == capec_id:  # Trova la CAPEC nei similarity_scores
                            model_rank = score_data[1].get("rank")
                            break
                    rank_data.append(model_rank)
                    print(f"{model}_Rank for CAPEC {capec_id}: {model_rank}")

                # Aggiungi la riga
                rows.append((cve_id, capec_id, capec_name, *rank_data))

    print(f"Finished processing SingleCorrelations. Total rows to write: {len(rows)}")

    # Ordina i dati nel foglio "results" per CVE_ID
    print("Sorting rows by CVE_ID...")
    rows = sorted(rows, key=lambda row: extract_cve_sort_key(row[0]))
    print("Rows sorted.")

    # Scrive i dati nel foglio "results"
    print("Writing results to the 'results' sheet...")
    for row in rows:
        print(f"Writing row: {row}")
        results_sheet.append(row)

    print(f"Results Rows Written: {len(rows)}")

    # Aggiunge i collegamenti ipertestuali
    print("Applying hyperlinks to 'results' sheet...")
    apply_hyperlinks(results_sheet, top_count=10)
    print("Hyperlinks applied.")

    # Trova gli indici delle colonne di ranking per la colorazione
    ranking_columns = [headers.index(f"{model}_Rank") + 1 for model in ai_models]  # Indici delle colonne (1-based)
    print(f"Ranking Columns: {ranking_columns}")

    # Applica la colorazione alle colonne di ranking
    print("Applying ranking colors...")
    apply_ranking_colors(results_sheet, ranking_columns, top_score=top_count)
    print("Ranking colors applied.")

    # Applica stili alle colonne di base (CVE_ID, CAPEC_ID, CAPEC_Name)
    print("Styling basic columns...")
    style_basic_columns(results_sheet, rows)
    print("Basic columns styled.")

    # Popola il foglio "stats" con le metriche
    print("Writing statistics to the 'stats' sheet...")
    stats_sheet.append(["Statistic", "Value"])  # Intestazione esempio
    total_cve_processed = len(groundtruth_mapping)
    matching_cve = len(set(row[0] for row in rows))
    total_capec_mappings = len(rows)

    print(f"Total CVE Processed: {total_cve_processed}")
    print(f"Matching CVE in SingleCorrelations: {matching_cve}")
    print(f"Total CAPEC Mappings: {total_capec_mappings}")

    stats_sheet.append(["Total CVE Processed", total_cve_processed])
    stats_sheet.append(["Matching CVE in SingleCorrelations", matching_cve])
    stats_sheet.append(["Total CAPEC Mappings", total_capec_mappings])

    # Calcolo dell'MRR e delle metriche di Recall per ciascun modello e aggiunta ai risultati
    stats_sheet.append(["Method", "MRR", "Recall@1", "Recall@5", "Recall@10", "Recall@20"])  # Intestazione

    for model in ai_models:
        # Estrae i rank di ogni CAPEC per il modello corrente
        model_ranks = [row[headers.index(f"{model}_Rank")] for row in rows if row[headers.index(f"{model}_Rank")] is not None]
        
        # Calcola l'MRR usando la funzione importata
        mrr_value = mean_reciprocal_rank(model_ranks)
        
        # Calcola le metriche di Recall
        recall_1 = recall_at_k(model_ranks, 1)
        recall_5 = recall_at_k(model_ranks, 5)
        recall_10 = recall_at_k(model_ranks, 10)
        recall_20 = recall_at_k(model_ranks, 20)
        
        # Aggiunge i risultati alla riga per il modello corrente
        stats_sheet.append([model, mrr_value, recall_1, recall_5, recall_10, recall_20])
        print(f"MRR for {model}: {mrr_value}, Recall@1: {recall_1}, Recall@5: {recall_5}, Recall@10: {recall_10}, Recall@20: {recall_20}")


    # Salva il workbook
    try:
        file_name = f"threatlinker_groundtruth_task_{task.id}_groundtruth_{groundtruth.id}.xlsx"
        print(f"Saving workbook to: {REPORTS / file_name}")
        save_excel_workbook(workbook, file_name, REPORTS)
        print(f"Workbook successfully saved: {REPORTS / file_name}")
        return REPORTS / file_name
    except Exception as e:
        print("Error saving workbook:")
        print(traceback.format_exc())
        raise e

def style_basic_columns(ws, rows):
    """
    Applica stili alle prime tre colonne (CVE_ID, CAPEC_ID, CAPEC_Name) di un foglio Excel.

    :param ws: Foglio Excel (worksheet) su cui applicare gli stili.
    :param rows: Righe da stilizzare.
    """
    # Definizione del bordo sottile
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Definizione dei colori per le colonne specifiche
    cve_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Grigio chiaro
    capec_fill = PatternFill(start_color="F5DEB3", end_color="F5DEB3", fill_type="solid")  # Beige chiaro

    # Itera sulle righe e applica i colori alle prime tre colonne
    for row_idx, row in enumerate(rows, start=2):  # Ignora l'intestazione (prima riga)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border  # Applica sempre il bordo

            # Applica colori solo alle prime tre colonne
            if col_idx == 1:  # CVE_ID
                cell.fill = cve_fill
            elif col_idx in [2, 3]:  # CAPEC_ID e CAPEC_Name
                cell.fill = capec_fill
