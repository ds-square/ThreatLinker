from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from pathlib import Path
import re 

def create_empty_excel_with_sheets(sheet_names):
    """
    Crea un file Excel vuoto con un certo numero di fogli specificati.

    :param sheet_names: Lista di nomi dei fogli da creare.
    :return: Una tupla contenente:
             - workbook (il file Excel creato).
             - sheet_objects (un dizionario con i nomi dei fogli come chiavi e i fogli come valori).
    """
    # Crea un nuovo workbook
    wb = Workbook()

    # Dizionario per restituire i fogli
    sheet_objects = {}

    # Crea i fogli
    for idx, sheet_name in enumerate(sheet_names):
        if idx == 0:
            # Usa il foglio di default per il primo
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Aggiungi il foglio al dizionario
        sheet_objects[sheet_name] = ws

    return wb, sheet_objects

def save_excel_workbook(workbook, file_name, destination):
    """
    Salva un workbook Excel nella directory specificata.
    Crea la directory di destinazione se non esiste.

    :param workbook: Workbook da salvare (oggetto openpyxl Workbook).
    :param file_name: Nome del file (incluso .xlsx).
    :param destination: Percorso della directory di destinazione (oggetto Path o stringa).
    """
    # Assicurati che la directory di destinazione sia un oggetto Path
    destination = Path(destination)

    # Crea la directory se non esiste
    if not destination.exists():
        destination.mkdir(parents=True, exist_ok=True)

    # Componi il percorso completo del file
    file_path = destination / file_name

    # Salva il workbook
    workbook.save(file_path)
    print(f"Workbook salvato con successo in: {file_path}")

def extract_cve_sort_key(cve_id):
    """
    Estrae una chiave di ordinamento dagli ID delle CVE nel formato CVE-XXXX-YYYY.

    :param cve_id: ID della CVE (es. "CVE-2021-12345").
    :return: Una tupla (XXXX, YYYY) per l'ordinamento.
    """
    match = re.match(r"CVE-(\d{4})-(\d+)", cve_id)
    if match:
        year, number = match.groups()
        return int(year), int(number)
    return float('inf'), float('inf')  # Valore alto per posizionare eventuali CVE non valide alla fine

def apply_hyperlinks(ws, top_count):
    """
    Aggiunge hyperlink ai CVE-ID e CAPEC-ID nelle colonne specifiche del foglio Excel.

    :param ws: Foglio Excel (worksheet) su cui applicare i collegamenti ipertestuali.
    :param top_count: Numero massimo di CAPEC da considerare.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Hyperlink per CVE-ID nella colonna 1 (prima colonna)
        cve_cell = ws.cell(row=row_idx, column=1)
        if cve_cell.value:
            cve_id = cve_cell.value
            cve_cell.hyperlink = f"https://nvd.nist.gov/vuln/detail/{cve_id}"
            cve_cell.style = "Hyperlink"

        # Hyperlink per CAPEC-ID nelle colonne Rank_X
        for capec_col in range(2, 2 + 2 * top_count, 2):  # Salta di 2: Rank_X
            capec_cell = ws.cell(row=row_idx, column=capec_col)
            if capec_cell.value and isinstance(capec_cell.value, str) and capec_cell.value.startswith("CAPEC-"):
                capec_id = capec_cell.value.split("-")[-1]  # Estrai l'ID numerico
                capec_cell.hyperlink = f"https://capec.mitre.org/data/definitions/{capec_id}.html"
                capec_cell.style = "Hyperlink"

def apply_ranking_colors(ws, ranking_columns, top_score):
    """
    Applica colori alle colonne di ranking in base ai valori dei numeri.

    :param ws: Foglio Excel (worksheet) su cui applicare i colori.
    :param ranking_columns: Lista di indici delle colonne di ranking da colorare.
    :param top_score: Numero massimo considerato "top score" per il verde chiaro.
    """
    # Definizione dei colori
    green_dark = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde scuro
    green_light = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde chiaro
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Giallo
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Arancione
    red = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid")  # Rosso

    # Itera su tutte le righe, partendo dalla seconda (ignorando l'intestazione)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for col_idx in ranking_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            if isinstance(value, int):  # Applica colore solo ai numeri interi
                if value == 1:
                    cell.fill = green_dark
                elif 1 < value <= top_score:
                    cell.fill = green_light
                elif top_score < value <= 30:
                    cell.fill = yellow
                elif 30 < value <= 50:
                    cell.fill = orange
                elif value > 50:
                    cell.fill = red