from django.shortcuts import get_object_or_404
from core.models import Task
from core.generator.generate_excel_utils import create_empty_excel_with_sheets, save_excel_workbook, extract_cve_sort_key, apply_hyperlinks
from threatlinker.paths import REPORTS

from openpyxl.styles import PatternFill, Font


import ast  # Per deserializzare stringhe rappresentanti strutture di dati Python


def create_excel_with_task_hosts(task_id, top_count, valid_models):
    """
    Crea un file Excel con un foglio per ogni combinazione di host e modello AI valido.
    Popola ogni foglio con i CVE associati all'host e i dati del modello.

    :param task_id: ID della Task da elaborare.
    :param top_count: Numero massimo di CAPECs da considerare.
    :param valid_models: Lista dei modelli AI validi selezionati.
    :return: Percorso del file Excel generato.
    """
    print(f"Starting function for Task ID: {task_id}")
    print(f"Top Count: {top_count}")
    print(f"Valid Models: {valid_models}")

    # Recupera la Task specifica
    task = get_object_or_404(Task, id=task_id)
    print(f"Task retrieved: {task}")

    # Estrai tutti gli host univoci da cve_hosts
    cve_hosts = task.cve_hosts  # Presuppone che cve_hosts sia un dizionario {CVE: [hosts]}
    print(f"Task.cve_hosts: {cve_hosts}")

    unique_hosts = set()
    for hosts in cve_hosts.values():
        # Se hosts è una stringa che rappresenta una lista, deserializzala
        if isinstance(hosts, str):
            try:
                hosts = ast.literal_eval(hosts)  # Converte la stringa in una lista
            except (ValueError, SyntaxError):
                print(f"Error: Unable to parse hosts: {hosts}")
                hosts = []
        unique_hosts.update(hosts)
    print(f"Unique Hosts: {unique_hosts}")

    # Recupera i modelli AI dalla Task e filtra per i modelli validi
    ai_models = task.ai_models if task.ai_models else []
    print(f"Task AI Models: {ai_models}")

    filtered_models = [model for model in ai_models if model in valid_models]
    print(f"Filtered Models: {filtered_models}")

    if not filtered_models:
        raise ValueError("No valid AI models selected for the Task.")

    # Crea i nomi per i fogli combinando host e modello AI valido
    sheet_names = [f"{host}-{model}" for host in sorted(unique_hosts) for model in filtered_models]
    print(f"Sheet Names: {sheet_names}")

    # Usa la funzione per creare il workbook
    workbook, sheet_objects = create_empty_excel_with_sheets(sheet_names)
    print(f"Workbook and Sheets created. Sheets: {list(sheet_objects.keys())}")

    # Dizionario per raccogliere i dati per ogni foglio
    sheet_data = {sheet_name: [] for sheet_name in sheet_names}

    # Popola i dati nei fogli
    for single_correlation in task.single_correlations.all():
        print(f"Processing SingleCorrelation: {single_correlation}")
        cve_id = single_correlation.cve_id

        # Assicurati che `hosts` sia una lista
        associated_hosts = single_correlation.hosts
        if isinstance(associated_hosts, str):
            try:
                associated_hosts = ast.literal_eval(associated_hosts)
            except (ValueError, SyntaxError):
                print(f"Error: Unable to parse associated_hosts: {associated_hosts}")
                associated_hosts = []

        print(f"CVE ID: {cve_id}, Associated Hosts: {associated_hosts}")

        for host in associated_hosts:
            print(f"Processing Host: {host}")
            for model in filtered_models:
                sheet_name = f"{host}-{model}"
                print(f"Processing Sheet: {sheet_name}")

                if sheet_name in sheet_data:
                    similarity_scores = single_correlation.similarity_scores.get(model, [])
                    print(f"Similarity Scores for Model {model}: {similarity_scores}")

                    # Ordina le CAPEC per rank
                    sorted_scores = sorted(similarity_scores, key=lambda x: x[1]["rank"])[:top_count]

                    # Costruisci la riga
                    row = [cve_id]  # Inizia con il CVE_ID
                    for capec, data in sorted_scores:
                        row.append(capec)  # Rank_X: CAPEC ID
                        row.append(data["final_score"])  # Score_X: Final Score

                    # Riempie i valori mancanti con None fino a top_count
                    while len(row) < (2 * top_count) + 1:
                        row.append(None)

                    sheet_data[sheet_name].append(row)

    # Ordina i dati nei fogli per CVE_ID
    for sheet_name, rows in sheet_data.items():
        print(f"Sorting rows for sheet: {sheet_name}")
        sheet_data[sheet_name] = sorted(rows, key=lambda row: extract_cve_sort_key(row[0]))
        print(f"Sorted rows for sheet: {sheet_name}")

    # Scrive i dati nei fogli
    for sheet_name, rows in sheet_data.items():
        ws = sheet_objects[sheet_name]
        print(f"Writing data to sheet: {sheet_name}")

        # Scrive l'intestazione
        headers = ["CVE_ID"] + [item for x in range(1, top_count + 1) for item in (f"Rank_{x}", f"Score_{x}")]
        ws.append(headers)
        print(f"Headers added to {sheet_name}: {headers}")

        # Scrive le righe
        for row in rows:
            ws.append(row)
            print(f"Row added to {sheet_name}: {row}")
        
        # Aggiunge hyperlink
        apply_hyperlinks(ws, top_count)
        
        # Applica stile al foglio
        apply_sheet_styles(ws, top_count)

    # Nome del file
    file_name = f"threatlinker_task_{task.id}_top_capecs.xlsx"
    print(f"File Name: {file_name}")

    # Salva il workbook nella directory 'reports'
    save_excel_workbook(workbook, file_name, REPORTS)
    print(f"Workbook saved in {REPORTS}")

    return REPORTS / file_name


def apply_sheet_styles(ws, top_count):
    """
    Applica stile agli header e alle colonne di un foglio Excel.
    Colora gli header e ogni coppia di colonne (Rank_X, Score_X) con lo stesso colore.

    :param ws: Foglio Excel (worksheet) su cui applicare lo stile.
    :param top_count: Numero massimo di coppie Rank_X e Score_X.
    """
    # Colori per header e coppie di colonne
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    pair_colors = ["FCE5CD", "D9EAD3", "CFE2F3", "FFF2CC"]  # Ciclo di colori per le coppie
    header_font = Font(bold=True)

    # Applica stile agli header
    for col_num, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font

    # Applica stile alle coppie di colonne Rank_X e Score_X
    for i in range(1, top_count + 1):
        rank_col = (i * 2)  # Colonna di Rank_X (es. 2, 4, 6, ...)
        score_col = (i * 2) + 1  # Colonna di Score_X (es. 3, 5, 7, ...)
        color = pair_colors[(i - 1) % len(pair_colors)]  # Ruota i colori

        rank_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        score_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for row in ws.iter_rows(min_col=rank_col, max_col=rank_col, min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.fill = rank_fill

        for row in ws.iter_rows(min_col=score_col, max_col=score_col, min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.fill = score_fill
